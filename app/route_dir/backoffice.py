from flask import Blueprint, render_template, redirect, session,abort, current_app, make_response, send_file
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
import uuid
import hashlib
import numpy
import os
import datetime
from config import config
from unidecode import unidecode


from ..model_dir.intervention import Intervention, InterventionValues
from ..model_dir.type_intervention import TypeIntervention, TypeInterventionSite

from ..model_dir.photo import Photo
from ..model_dir.place import Place
from ..model_dir.form import Form
from ..model_dir.field import Field
from ..model_dir.field_histo import FieldHisto
from ..model_dir.mymixin import User
from ..model_dir.company import Company

from flask import jsonify, request, abort, send_from_directory
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.utils import secure_filename

from .. import db,  getByIdOrByName, getByIdOrFilename
app_file_backoffice= Blueprint('backoffice',__name__)

import cv2
import json
from sqlalchemy import desc, text

# Setup upload folder
UPLOAD_FOLDER = config["upload_dir"]

class XLWrap:
    def __init__(self, wb, filename):
        self.wb = wb
        self.filename = filename

    def __getitem__(self, key):
        # Returns the value for cells given set name
        return [
            self.wb[sheet][cell].value
            for sheet, cell in list(self.wb.defined_names[key].destinations)
        ]

    def __setitem__(self, key, dat):
        # Sets the value for cells in workbook given set name
        cells = self.wb.defined_names[key].destinations
        for sheet, cell in cells:
            ws = self.wb[sheet]
            ws[cell] = dat
        self.save()

    def save(self):
        self.wb.save(self.filename)
        

@app_file_backoffice.route("/intervention_values/feb/xlsx/<id>", methods=["GET"])
def get_interventions_values_id(id):
    _intervention_values = InterventionValues.query.get(id)
    if _intervention_values is None:
        abort(make_response(jsonify(error="intervention_values is not found"), 404))
        
    _type_intervention_site = TypeInterventionSite.query.get((_intervention_values.type_intervention.id,_intervention_values.site.id))
    if _type_intervention_site is None:
        abort(make_response(jsonify(error="_type_intervention_site is not found"), 404))

    filename_input='/Users/ysimonx/Developpement/on_site_intervention_api/app/static/assets/scaffolding_request_sandbox_input.xlsx'
    print(filename_input)
    filename_output = '/tmp/scaffolding_request_sandbox_output_{}.xlsx'.format(_intervention_values.name)

    dict_column_values = _intervention_values.to_dict()
    dict_converted={}
    
    dict_values = dict_column_values["data"]
    
    # conversion sans accents des colonnes et valeurs (excel aime pas)
    for key, value in dict_values.items():
        if value is not None and str(value).startswith("<svg"):
            value="yes"
        if value is None:
            value=""
        column=key.strip()
        column = column.replace("-","_")
        column = column.replace(" ","_")
        column = unidecode(str(column))
        dict_converted[column]=unidecode(str(value))
            

    wb = load_workbook(filename_input)
    cells = wb.defined_names
    
    xlw = XLWrap(wb, filename_output)
    # xlw["num_registre"] = _intervention_values.name
    for cell_name, values in cells.items():
        
        print(cell_name)
        if cell_name in dict_converted:
            print("it's a MATCH for {}".format(cell_name))
            try:
                value =  dict_converted[cell_name]
                if len(value) > 0:
                    xlw[cell_name] = dict_converted[cell_name]
                else:
                    xlw[cell_name] = " "
                
            except:
                print("zarb")
        else:
            xlw[cell_name] = " " # valeur par defaut
            
    date = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S.%f")

    download_name="feb_{}_{}.xlsx".format(_intervention_values.name, date)
    return send_file(filename_output, download_name=download_name)
	
    # return render_template('feb.html', intervention_values=_intervention_values, type_intervention_site=_type_intervention_site)
   


